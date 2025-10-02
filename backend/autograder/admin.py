from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.admin import AdminSite
from django import forms
from accounts.models import User
from schools.models import School
from classrooms.models import Classroom, ClassroomPermission, MembershipType, SchoolBillingReport
from students.models import Student
from scores.models import Score, TestResult, CommentTemplate, TestSummary, SchoolTestSummary
from tests.models import TestSchedule, TestDefinition, QuestionGroup, Question, AnswerKey


def format_grade_display(grade):
    """学年を「小6」「中1」形式で表示する共通関数"""
    if not grade:
        return '未設定'
    
    try:
        grade_num = int(grade)
        if 1 <= grade_num <= 6:
            return f'小{grade_num}'
        elif 7 <= grade_num <= 9:
            return f'中{grade_num - 6}'
        else:
            return str(grade)
    except (ValueError, TypeError):
        return str(grade)


# 標準管理サイトをカスタマイズ
class CustomAdminSite(AdminSite):
    site_header = '全国学力向上テスト 管理サイト'
    site_title = '全国学力向上テスト'
    index_title = '管理画面'

    def get_app_list(self, request, app_label=None):
        """
        カスタムアプリリストを返す - ユーザー管理と塾管理を分離
        """
        app_dict = {}
        
        # ユーザー管理カテゴリを作成
        app_dict['ユーザー管理'] = {
            'name': 'ユーザー管理',
            'app_label': 'ユーザー管理',
            'app_url': None,
            'has_module_perms': True,
            'models': []
        }
        
        # 塾管理カテゴリを作成
        app_dict['塾管理'] = {
            'name': '塾管理',
            'app_label': '塾管理',
            'app_url': None,
            'has_module_perms': True,
            'models': []
        }
        
        # テスト問題カテゴリを作成
        app_dict['テスト問題'] = {
            'name': 'テスト問題',
            'app_label': 'テスト問題',
            'app_url': None,
            'has_module_perms': True,
            'models': []
        }
        
        # モデルをカテゴリ別に分類
        if self.has_permission(request):
            for model, model_admin in self._registry.items():
                app_label = model._meta.app_label
                
                # モデル情報を作成
                model_dict = {
                    'name': model._meta.verbose_name_plural,
                    'object_name': model.__name__,
                    'perms': {
                        'add': model_admin.has_add_permission(request),
                        'change': model_admin.has_change_permission(request),
                        'delete': model_admin.has_delete_permission(request),
                        'view': model_admin.has_view_permission(request),
                    },
                    'admin_url': None,
                    'add_url': None,
                }
                
                if model_dict['perms']['change'] or model_dict['perms']['view']:
                    model_dict['view_only'] = not model_dict['perms']['change']
                    try:
                        model_dict['admin_url'] = self.reverse(f'admin:{app_label}_{model.__name__.lower()}_changelist')
                    except:
                        pass
                
                if model_dict['perms']['add']:
                    try:
                        model_dict['add_url'] = self.reverse(f'admin:{app_label}_{model.__name__.lower()}_add')
                    except:
                        pass
                
                # モデルをカテゴリ別に分類
                if model.__name__ == 'User':
                    app_dict['ユーザー管理']['models'].append(model_dict)
                else:
                    app_dict['塾管理']['models'].append(model_dict)
        
        # 塾管理モデルを順序付け（塾、教室、生徒の順）
        juku_model_order = ['塾', '教室', '生徒']
        app_dict['塾管理']['models'].sort(key=lambda x: juku_model_order.index(x['name']) if x['name'] in juku_model_order else 999)
        
        # 結果リストを作成
        result = []
        if app_dict['ユーザー管理']['models']:
            result.append(app_dict['ユーザー管理'])
        if app_dict['塾管理']['models']:
            result.append(app_dict['塾管理'])
        
        return result


# 標準管理サイトをカスタマイズ
admin.site.site_header = 'アン小学生テスト 管理サイト'
admin.site.site_title = 'アン小学生テスト'
admin.site.index_title = '管理画面'

# 標準管理サイトのget_app_listをオーバーライド
original_get_app_list = admin.site.get_app_list

def custom_get_app_list(self, request, app_label=None):
    """
    カスタムアプリリストを返す - ユーザー管理と塾管理を分離
    """
    app_dict = {}
    
    # ユーザー管理カテゴリを作成
    app_dict['ユーザー管理'] = {
        'name': 'ユーザー管理',
        'app_label': 'ユーザー管理',
        'app_url': None,
        'has_module_perms': True,
        'models': []
    }
    
    # 塾管理カテゴリを作成
    app_dict['塾管理'] = {
        'name': '塾管理',
        'app_label': '塾管理',
        'app_url': None,
        'has_module_perms': True,
        'models': []
    }
    
    # テスト問題カテゴリを作成
    app_dict['テスト問題'] = {
        'name': 'テスト問題',
        'app_label': 'テスト問題',
        'app_url': None,
        'has_module_perms': True,
        'models': []
    }
    
    # モデルをカテゴリ別に分類
    if self.has_permission(request):
        for model, model_admin in self._registry.items():
            app_label_orig = model._meta.app_label
            
            # モデル情報を作成
            model_dict = {
                'name': model._meta.verbose_name_plural,
                'object_name': model.__name__,
                'perms': {
                    'add': model_admin.has_add_permission(request),
                    'change': model_admin.has_change_permission(request),
                    'delete': model_admin.has_delete_permission(request),
                    'view': model_admin.has_view_permission(request),
                },
                'admin_url': None,
                'add_url': None,
            }
            
            if model_dict['perms']['change'] or model_dict['perms']['view']:
                model_dict['view_only'] = not model_dict['perms']['change']
                try:
                    model_dict['admin_url'] = f'/admin/{app_label_orig}/{model.__name__.lower()}/'
                except:
                    pass
            
            if model_dict['perms']['add']:
                try:
                    model_dict['add_url'] = f'/admin/{app_label_orig}/{model.__name__.lower()}/add/'
                except:
                    pass
            
            # モデルをカテゴリ別に分類
            if model.__name__ == 'User':
                app_dict['ユーザー管理']['models'].append(model_dict)
            elif model.__name__ in ['School', 'Classroom', 'Student', 'Score', 'TestResult', 'CommentTemplate', 'MembershipType', 'SchoolBillingReport']:
                app_dict['塾管理']['models'].append(model_dict)
            elif model.__name__ in ['TestSchedule', 'TestDefinition', 'QuestionGroup', 'Question', 'AnswerKey']:
                app_dict['テスト問題']['models'].append(model_dict)
    
    # 各カテゴリのモデルを順序付け
    juku_model_order = ['塾', '会員種別', '教室', '生徒', '得点', 'テスト結果', 'コメントテンプレート', '塾別課金レポート']
    app_dict['塾管理']['models'].sort(key=lambda x: juku_model_order.index(x['name']) if x['name'] in juku_model_order else 999)
    
    test_model_order = ['テストスケジュール', 'テスト', '大問', '問題', '解答']
    if 'テスト問題' in app_dict:
        app_dict['テスト問題']['models'].sort(key=lambda x: test_model_order.index(x['name']) if x['name'] in test_model_order else 999)
    
    # 結果リストを作成
    result = []
    if app_dict['ユーザー管理']['models']:
        result.append(app_dict['ユーザー管理'])
    if app_dict['塾管理']['models']:
        result.append(app_dict['塾管理'])
    if app_dict['テスト問題']['models']:
        result.append(app_dict['テスト問題'])
    
    return result

# 標準管理サイトのメソッドを置き換え
admin.site.get_app_list = custom_get_app_list.__get__(admin.site, admin.site.__class__)


# ユーザー管理
class CustomUserAdmin(BaseUserAdmin):
    list_display = ('username', 'email', 'role', 'school_id', 'classroom_id', 'is_active', 'date_joined')
    list_filter = ('role', 'is_active', 'is_staff', 'is_superuser')
    search_fields = ('username', 'email', 'school_id', 'classroom_id')
    ordering = ('username',)
    
    fieldsets = BaseUserAdmin.fieldsets + (
        ('追加情報', {'fields': ('role', 'school_id', 'classroom_id')}),
    )
    
    add_fieldsets = BaseUserAdmin.add_fieldsets + (
        ('追加情報', {'fields': ('role', 'school_id', 'classroom_id')}),
    )


# 塾管理
class ClassroomInline(admin.TabularInline):
    model = Classroom
    extra = 0
    readonly_fields = ('created_at', 'updated_at')


class SchoolAdmin(admin.ModelAdmin):
    list_display = ('school_id', 'name', 'contact_person', 'membership_type', 'get_price_display', 'get_status_with_date', 'email', 'phone', 'can_access', 'created_at')
    list_filter = ('status', 'membership_type', 'is_active', 'created_at')
    search_fields = ('school_id', 'name', 'contact_person', 'email')
    readonly_fields = ('created_at', 'updated_at')
    inlines = [ClassroomInline]
    actions = ['export_template', 'export_existing_schools']
    
    def get_status_with_date(self, obj):
        return obj.get_status_display_with_date()
    get_status_with_date.short_description = 'ステータス'
    get_status_with_date.admin_order_field = 'status'
    
    def get_price_display(self, obj):
        return f"{obj.get_price_per_student()}円/名"
    get_price_display.short_description = '料金'
    get_price_display.admin_order_field = 'membership_type'
    
    def can_access(self, obj):
        return obj.can_access()
    can_access.boolean = True
    can_access.short_description = 'アクセス可能'
    
    fieldsets = (
        ('基本情報', {
            'fields': ('school_id', 'name', 'contact_person', 'email', 'phone', 'address', 'membership_type')
        }),
        ('ステータス管理', {
            'fields': ('status', 'trial_date', 'active_date', 'withdrawn_date', 'is_active'),
            'description': 'ステータス変更時に対応する日付が自動設定されます。'
        }),
        ('権限設定', {
            'fields': ('can_register_students', 'can_input_scores', 'can_view_reports'),
            'description': 'この塾の教室管理者が実行できる操作を設定します。'
        }),
        ('システム情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def export_template(self, request, queryset):
        """塾登録用テンプレートをダウンロード"""
        from django.http import HttpResponse
        from schools.utils import export_school_template
        import io
        import pandas as pd
        
        # 選択されたオブジェクトが不要なアクションであることを明示
        # テンプレートダウンロードは常に実行可能
        
        df = export_school_template()
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='塾登録テンプレート')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="school_template.xlsx"'
        return response
    
    # アクションが選択不要であることを明示
    export_template.short_description = "📝 塾登録テンプレートをダウンロード"
    export_template.allowed_permissions = ('add',)
    
    def export_existing_schools(self, request, queryset):
        """既存塾データをエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        # 選択された塾データまたは全データをエクスポート
        if queryset.exists():
            schools = queryset
        else:
            schools = self.get_queryset(request)
        
        data = []
        for school in schools:
            data.append({
                '塾ID': school.school_id,
                '塾名': school.name,
                '担当者名': school.contact_person,
                'メールアドレス': school.email,
                '電話番号': school.phone or '',
                '住所': school.address or '',
                'ステータス': school.get_status_display(),
                'ステータス日付': school.get_status_display_with_date(),
                '有効状態': '有効' if school.is_active else '無効',
                '登録日': school.created_at.strftime('%Y-%m-%d'),
                '教室数': school.classrooms.count(),
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # データシート
            df.to_excel(writer, index=False, sheet_name='塾データ')
            
            # テンプレートシート
            from schools.utils import export_school_template
            template_df = export_school_template()
            template_df.to_excel(writer, index=False, sheet_name='新規登録テンプレート')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="schools_export_{timestamp}.xlsx"'
        return response
    
    export_existing_schools.short_description = "📊 既存塾データをエクスポート"
    
    
    def get_actions(self, request):
        """利用可能なアクションを返す"""
        actions = super().get_actions(request)
        return actions
    
    def changelist_view(self, request, extra_context=None):
        """変更リストビューをカスタマイズ"""
        extra_context = extra_context or {}
        extra_context['import_url'] = '/admin/schools/import/'
        extra_context['template_download_url'] = '/download-school-template/'
        return super().changelist_view(request, extra_context)
    
    def get_urls(self):
        """カスタムURLを追加"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-schools/', self.admin_site.admin_view(self.import_schools_view), name='schools_school_import_direct'),
        ]
        return custom_urls + urls
    
    def import_schools_view(self, request):
        """直接インポートビュー（選択不要）"""
        from schools.admin_actions import school_import_action
        return school_import_action(self, request, None)


# 教室管理
class StudentInline(admin.TabularInline):
    model = Student
    extra = 0
    readonly_fields = ('created_at', 'updated_at')


class ClassroomPermissionInline(admin.StackedInline):
    model = ClassroomPermission
    extra = 0
    max_num = 1
    readonly_fields = ('created_at', 'updated_at')
    fieldsets = (
        ('権限設定', {
            'fields': ('can_register_students', 'can_input_scores', 'can_view_reports')
        }),
        ('システム情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


class ClassroomAdmin(admin.ModelAdmin):
    list_display = ('classroom_id', 'name', 'school', 'get_membership_type', 'get_price_display', 'is_active', 'created_at')
    list_filter = ('is_active', 'created_at', 'school')
    search_fields = ('classroom_id', 'name', 'school__name')
    readonly_fields = ('created_at', 'updated_at')
    inlines = [StudentInline, ClassroomPermissionInline]
    
    def get_membership_type(self, obj):
        return obj.school.get_membership_type_display()
    get_membership_type.short_description = '会員種別'
    get_membership_type.admin_order_field = 'school__membership_type'
    
    def get_price_display(self, obj):
        return f"{obj.get_price_per_student()}円/名"
    get_price_display.short_description = '料金'
    get_price_display.admin_order_field = 'school__membership_type'
    
    fieldsets = (
        ('基本情報', {
            'fields': ('classroom_id', 'name', 'school', 'is_active')
        }),
        ('システム情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


# 生徒管理
class StudentAdmin(admin.ModelAdmin):
    list_display = ('get_school_id', 'get_school_name', 'get_classroom_name', 'student_id', 'name', 'get_grade_display', 'is_active', 'created_at')
    list_filter = ('is_active', 'grade', 'created_at', 'classroom__school__school_id', 'classroom__school')
    search_fields = ('student_id', 'name', 'classroom__name', 'classroom__school__name', 'classroom__school__school_id')
    readonly_fields = ('created_at', 'updated_at', 'get_grade_display_readonly')
    ordering = ('classroom__school__school_id', 'classroom__classroom_id', 'student_id')
    actions = ['export_students', 'export_students_by_school', 'export_test_participants', 'export_all_test_participants']
    
    def get_school_id(self, obj):
        return obj.classroom.school.school_id
    get_school_id.short_description = '塾ID'
    get_school_id.admin_order_field = 'classroom__school__school_id'
    
    def get_school_name(self, obj):
        return obj.classroom.school.name
    get_school_name.short_description = '塾名'
    get_school_name.admin_order_field = 'classroom__school__name'
    
    def get_classroom_name(self, obj):
        return obj.classroom.name
    get_classroom_name.short_description = '教室名'
    get_classroom_name.admin_order_field = 'classroom__name'
    
    def get_grade_display(self, obj):
        """学年を「小6」「中1」形式で表示"""
        return format_grade_display(obj.grade)
    get_grade_display.short_description = '学年'
    get_grade_display.admin_order_field = 'grade'
    
    def get_grade_display_readonly(self, obj):
        """詳細画面用の読み取り専用学年表示"""
        return format_grade_display(obj.grade)
    get_grade_display_readonly.short_description = '学年表示'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('classroom', 'classroom__school')
    
    def export_students(self, request, queryset):
        """選択した生徒データをエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        # エクスポートデータを準備
        data = []
        for student in queryset:
            data.append({
                '塾ID': student.classroom.school.school_id,
                '塾名': student.classroom.school.name,
                '教室ID': student.classroom.classroom_id,
                '教室名': student.classroom.name,
                '生徒ID': student.student_id,
                '生徒名': student.name,
                '学年': format_grade_display(student.grade),
                '有効状態': '有効' if student.is_active else '無効',
                '登録日': student.created_at.strftime('%Y-%m-%d'),
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='生徒一覧')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="students_export_{timestamp}.xlsx"'
        return response
    
    export_students.short_description = "選択した生徒データをエクスポート"
    
    def export_students_by_school(self, request, queryset):
        """塾別に生徒データをエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        # 塾別にグループ化
        schools = {}
        for student in queryset:
            school_id = student.classroom.school.school_id
            if school_id not in schools:
                schools[school_id] = {
                    'name': student.classroom.school.name,
                    'students': []
                }
            
            schools[school_id]['students'].append({
                '教室ID': student.classroom.classroom_id,
                '教室名': student.classroom.name,
                '生徒ID': student.student_id,
                '生徒名': student.name,
                '学年': format_grade_display(student.grade),
                '有効状態': '有効' if student.is_active else '無効',
                '登録日': student.created_at.strftime('%Y-%m-%d'),
            })
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for school_id, school_data in schools.items():
                df = pd.DataFrame(school_data['students'])
                sheet_name = f"{school_id}_{school_data['name'][:10]}"  # シート名の長さ制限
                df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="students_by_school_{timestamp}.xlsx"'
        return response
    
    export_students_by_school.short_description = "塾別に生徒データをエクスポート"
    
    def export_test_participants(self, request, queryset):
        """テスト受講生徒一覧をエクスポート（年度・期間選択対応）"""
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from django.http import HttpResponse
        from students.models import StudentEnrollment
        from classrooms.models import AttendanceRecord
        import io
        import pandas as pd
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill
        
        # POSTリクエストの場合、エクスポート実行
        if request.method == 'POST':
            year = request.POST.get('year')
            period = request.POST.get('period')
            
            if not year or not period:
                messages.error(request, '年度と期間を選択してください。')
                return redirect(request.get_full_path())
            
            try:
                # 指定された年度・期間の受講生徒を取得
                enrollments = StudentEnrollment.objects.filter(
                    year=year,
                    period=period,
                    is_active=True
                ).select_related(
                    'student',
                    'student__classroom',
                    'student__classroom__school'
                ).order_by(
                    'student__classroom__school__school_id',
                    'student__classroom__classroom_id',
                    'student__student_id'
                )
                
                # データをExcel用に整理（出欠情報含む）
                data = []
                for enrollment in enrollments:
                    student = enrollment.student
                    classroom = student.classroom
                    school = classroom.school
                    
                    # 出欠記録を取得
                    attendance_records = AttendanceRecord.objects.filter(
                        classroom=classroom,
                        student_id=student.student_id,
                        year=year,
                        period=period
                    )
                    
                    # 科目別出欠状況を集計
                    subjects_attendance = {}
                    total_subjects = 0
                    attended_subjects = 0
                    
                    for record in attendance_records:
                        subjects_attendance[record.subject] = '出席' if record.has_score_input else '欠席'
                        total_subjects += 1
                        if record.has_score_input:
                            attended_subjects += 1
                    
                    attendance_rate = f"{attended_subjects}/{total_subjects}" if total_subjects > 0 else "0/0"
                    attendance_percentage = f"{(attended_subjects/total_subjects*100):.1f}%" if total_subjects > 0 else "0.0%"
                    
                    data.append({
                        '塾ID': school.school_id,
                        '塾名': school.name,
                        '教室名': classroom.name,
                        '生徒ID': student.student_id,
                        '生徒名': student.name,
                        '学年': format_grade_display(student.grade),
                        '年度': enrollment.year,
                        '期間': enrollment.get_period_display(),
                        '受講開始日': enrollment.enrolled_at.strftime('%Y-%m-%d'),
                        '出席率': attendance_rate,
                        '出席割合': attendance_percentage,
                        '科目別出欠': ', '.join([f"{subj}:{status}" for subj, status in subjects_attendance.items()]) if subjects_attendance else '記録なし',
                        'アクティブ': '有効' if student.is_active else '無効'
                    })
                
                if not data:
                    messages.error(request, '指定された条件の受講生徒が見つかりません。')
                    return redirect(request.get_full_path())
                
                # DataFrameを作成してExcelファイルを生成
                df = pd.DataFrame(data)
                
                output = io.BytesIO()
                
                # Excelライターを使用してフォーマットを設定
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='テスト受講生徒一覧', index=False)
                    
                    # ワークシートを取得
                    worksheet = writer.sheets['テスト受講生徒一覧']
                    
                    # 列幅を自動調整
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # ヘッダー行のスタイルを設定
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                
                output.seek(0)
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                
                filename = f'test_participants_{year}_{period}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
            except Exception as e:
                messages.error(request, f'エクスポートエラー: {str(e)}')
                return redirect(request.get_full_path())
        
        # GETリクエストの場合、年度・期間選択画面を表示
        years = [
            {'value': '2025', 'label': '2025年度'},
            {'value': '2026', 'label': '2026年度'},
            {'value': '2027', 'label': '2027年度'},
        ]
        
        periods = [
            {'value': 'spring', 'label': '春期'},
            {'value': 'summer', 'label': '夏期'},
            {'value': 'winter', 'label': '冬期'},
        ]
        
        context = {
            'title': 'テスト受講生徒一覧エクスポート',
            'opts': self.model._meta,
            'years': years,
            'periods': periods,
            'action_name': 'export_test_participants',
        }
        
        return render(request, 'admin/export_test_participants.html', context)
    
    export_test_participants.short_description = "📋 テスト受講生徒一覧をエクスポート"
    
    def export_all_test_participants(self, request, queryset):
        """全テスト受講生徒一覧を一括エクスポート（年度・期間別シート）"""
        from django.http import HttpResponse
        from students.models import StudentEnrollment
        from classrooms.models import AttendanceRecord
        import io
        import pandas as pd
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill
        
        try:
            # 全ての受講生徒を年度・期間別に取得
            enrollments = StudentEnrollment.objects.filter(
                is_active=True
            ).select_related(
                'student',
                'student__classroom',
                'student__classroom__school'
            ).order_by(
                'year',
                'period',
                'student__classroom__school__school_id',
                'student__classroom__classroom_id',
                'student__student_id'
            )
            
            if not enrollments.exists():
                self.message_user(request, '受講生徒データが見つかりません。', level='error')
                return
            
            # 年度・期間別にグループ化
            grouped_data = {}
            for enrollment in enrollments:
                key = f"{enrollment.year}_{enrollment.period}"
                if key not in grouped_data:
                    grouped_data[key] = []
                
                student = enrollment.student
                classroom = student.classroom
                school = classroom.school
                
                # 出欠記録を取得
                attendance_records = AttendanceRecord.objects.filter(
                    classroom=classroom,
                    student_id=student.student_id,
                    year=enrollment.year,
                    period=enrollment.period
                )
                
                # 科目別出欠状況を集計
                subjects_attendance = {}
                total_subjects = 0
                attended_subjects = 0
                
                for record in attendance_records:
                    subjects_attendance[record.subject] = '出席' if record.has_score_input else '欠席'
                    total_subjects += 1
                    if record.has_score_input:
                        attended_subjects += 1
                
                attendance_rate = f"{attended_subjects}/{total_subjects}" if total_subjects > 0 else "0/0"
                attendance_percentage = f"{(attended_subjects/total_subjects*100):.1f}%" if total_subjects > 0 else "0.0%"
                
                grouped_data[key].append({
                    '塾ID': school.school_id,
                    '塾名': school.name,
                    '教室名': classroom.name,
                    '生徒ID': student.student_id,
                    '生徒名': student.name,
                    '学年': format_grade_display(student.grade),
                    '年度': enrollment.year,
                    '期間': enrollment.get_period_display(),
                    '受講開始日': enrollment.enrolled_at.strftime('%Y-%m-%d'),
                    '出席率': attendance_rate,
                    '出席割合': attendance_percentage,
                    '科目別出欠': ', '.join([f"{subj}:{status}" for subj, status in subjects_attendance.items()]) if subjects_attendance else '記録なし',
                    'アクティブ': '有効' if student.is_active else '無効'
                })
            
            # Excelファイルを生成（年度・期間別にシート分け）
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 統計情報シートを作成
                summary_data = []
                for key, data_list in grouped_data.items():
                    year, period = key.split('_')
                    period_display = {'spring': '春期', 'summer': '夏期', 'winter': '冬期'}.get(period, period)
                    
                    total_students = len(data_list)
                    active_students = len([d for d in data_list if d['アクティブ'] == '有効'])
                    schools_count = len(set([d['塾ID'] for d in data_list]))
                    classrooms_count = len(set([f"{d['塾ID']}-{d['教室名']}" for d in data_list]))
                    
                    summary_data.append({
                        '年度': year,
                        '期間': period_display,
                        '総生徒数': total_students,
                        '有効生徒数': active_students,
                        '塾数': schools_count,
                        '教室数': classrooms_count
                    })
                    
                    # 年度・期間別シートを作成
                    df = pd.DataFrame(data_list)
                    sheet_name = f"{year}年度_{period_display}"
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # シートのスタイルを設定
                    worksheet = writer.sheets[sheet_name]
                    
                    # 列幅を自動調整
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # ヘッダー行のスタイルを設定
                    header_font = Font(bold=True)
                    header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                
                # 統計情報シートを最初に追加
                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='統計情報', index=False)
                    
                    # 統計シートのスタイルを設定
                    summary_worksheet = writer.sheets['統計情報']
                    for column in summary_worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 20)
                        summary_worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # ヘッダー行のスタイルを設定
                    header_font = Font(bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    
                    for cell in summary_worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
            
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'all_test_participants_{timestamp}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            self.message_user(request, f'全テスト受講生徒データをエクスポートしました（{len(grouped_data)}期間、{sum(len(data) for data in grouped_data.values())}名）')
            return response
            
        except Exception as e:
            self.message_user(request, f'エクスポートエラー: {str(e)}', level='error')
    
    export_all_test_participants.short_description = "📊 全テスト受講生徒を一括エクスポート（年度・期間別シート）"
    
    def get_urls(self):
        """カスタムURLを追加"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-students/', self.admin_site.admin_view(self.import_students_view), name='students_student_import'),
            path('export-students/', self.admin_site.admin_view(self.export_students_view), name='students_student_export'),
        ]
        return custom_urls + urls
    
    def import_students_view(self, request):
        """生徒インポートビュー"""
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from students.utils import import_students_by_school_from_excel, export_students_by_school_template
        from django.http import HttpResponse
        import tempfile
        import pandas as pd
        
        if request.method == 'POST':
            if 'excel_file' in request.FILES:
                file = request.FILES['excel_file']
                
                # 一時ファイルに保存
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    for chunk in file.chunks():
                        tmp_file.write(chunk)
                    tmp_file_path = tmp_file.name
                
                try:
                    # インポート実行（全ての塾から）
                    result = import_students_by_school_from_excel(tmp_file_path, None)
                    
                    if result['success']:
                        success_msg = f"生徒データのインポートが完了しました。作成された生徒: {result['created_students']}件"
                        messages.success(request, success_msg)
                        
                        if result['errors']:
                            error_msg = "以下のエラーがありました:\n" + "\n".join(result['errors'])
                            messages.warning(request, error_msg)
                    else:
                        messages.error(request, f"インポートに失敗しました: {result['error']}")
                        
                finally:
                    import os
                    os.unlink(tmp_file_path)
                
                return redirect('/admin/students/student/')
            
            elif 'download_template' in request.POST:
                # テンプレートダウンロード
                df = export_students_by_school_template()
                
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    df.to_excel(tmp_file.name, index=False)
                    tmp_file_path = tmp_file.name
                
                try:
                    with open(tmp_file_path, 'rb') as f:
                        response = HttpResponse(
                            f.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                        )
                        response['Content-Disposition'] = 'attachment; filename="student_template.xlsx"'
                        return response
                finally:
                    import os
                    os.unlink(tmp_file_path)
        
        context = {
            'title': '生徒データ一括インポート',
            'opts': Student._meta,
        }
        return render(request, 'admin/student_import.html', context)
    
    def export_students_view(self, request):
        """生徒エクスポートビュー（包括的データ）"""
        from django.http import HttpResponse
        from django.shortcuts import render
        from django.db.models import Q, Prefetch
        import io
        import pandas as pd
        from datetime import datetime
        from scores.models import Score, TestResult, TestAttendance
        from tests.models import TestDefinition, QuestionGroup, TestSchedule
        
        if request.method == 'POST':
            year = request.POST.get('year')
            period = request.POST.get('period')
            
            # テストフィルター
            test_filters = Q(is_active=True)
            if year and year != 'all':
                test_filters &= Q(schedule__year=int(year))
            if period and period != 'all':
                test_filters &= Q(schedule__period=period)
            
            # 生徒データを効率的に取得
            students = Student.objects.select_related(
                'classroom__school'
            ).prefetch_related(
                Prefetch(
                    'scores',
                    queryset=Score.objects.select_related(
                        'test__schedule',
                        'question_group'
                    ).filter(test__in=TestDefinition.objects.filter(test_filters))
                ),
                Prefetch(
                    'test_results',
                    queryset=TestResult.objects.select_related(
                        'test__schedule'
                    ).filter(test__in=TestDefinition.objects.filter(test_filters))
                ),
                Prefetch(
                    'test_attendances',
                    queryset=TestAttendance.objects.select_related(
                        'test__schedule'
                    ).filter(test__in=TestDefinition.objects.filter(test_filters))
                )
            ).filter(is_active=True)
            
            export_data = []
            
            for student in students:
                # テストごとにデータを整理
                test_scores = {}
                test_results = {}
                test_attendances = {}
                
                for score in student.scores.all():
                    test_id = score.test.id
                    if test_id not in test_scores:
                        test_scores[test_id] = {}
                    test_scores[test_id][score.question_group.group_number] = score
                
                for result in student.test_results.all():
                    test_results[result.test.id] = result
                
                for attendance in student.test_attendances.all():
                    test_attendances[attendance.test.id] = attendance
                
                # 各テストのデータを生成
                for test_id, scores_by_group in test_scores.items():
                    if scores_by_group:  # スコアがある場合のみ
                        first_score = list(scores_by_group.values())[0]
                        test = first_score.test
                        
                        row_data = {
                            '塾ID': student.classroom.school.school_id,
                            '塾名': student.classroom.school.name,
                            '教室ID': student.classroom.classroom_id,
                            '教室名': student.classroom.name,
                            '生徒ID': student.student_id,
                            '生徒名': student.name,
                            '学年': format_grade_display(student.grade),
                            '年度': test.schedule.year,
                            '期間': test.schedule.get_period_display(),
                            '科目名': test.get_subject_display(),
                        }
                        
                        # 出席情報
                        if test_id in test_attendances:
                            attendance = test_attendances[test_id]
                            row_data['出席状態'] = attendance.get_attendance_status_display()
                            row_data['出席済み'] = 'はい' if attendance.is_present else 'いいえ'
                        else:
                            # Scoreモデルのattendanceフィールドから取得
                            row_data['出席状態'] = '出席' if first_score.attendance else '欠席'
                            row_data['出席済み'] = 'はい' if first_score.attendance else 'いいえ'
                        
                        # 大問ごとの点数（最大10問まで対応）
                        total_score = 0
                        max_group_num = max(scores_by_group.keys()) if scores_by_group else 0
                        
                        for i in range(1, 11):  # 大問1～10
                            if i in scores_by_group:
                                score_value = scores_by_group[i].score
                                row_data[f'大問{i}'] = score_value
                                total_score += score_value
                            else:
                                row_data[f'大問{i}'] = '-'
                        
                        # 合計点数
                        if test_id in test_results:
                            row_data['合計点数'] = test_results[test_id].total_score
                            row_data['正答率'] = f"{float(test_results[test_id].correct_rate):.1f}%"
                        else:
                            row_data['合計点数'] = total_score
                            row_data['正答率'] = '-'
                        
                        export_data.append(row_data)
            
            # データがない場合の処理
            if not export_data:
                # 空のデータでも基本構造を返す
                export_data = [{
                    '塾ID': '',
                    '塾名': '',
                    '教室ID': '',
                    '教室名': '',
                    '生徒ID': '',
                    '生徒名': '',
                    '学年': '',
                    '年度': '',
                    '期間': '',
                    '科目名': '',
                    '出席状態': '',
                    '出席済み': '',
                    **{f'大問{i}': '' for i in range(1, 11)},
                    '合計点数': '',
                    '正答率': '',
                }]
            
            df = pd.DataFrame(export_data)
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='生徒テストデータ')
                
                # スタイリング
                worksheet = writer.sheets['生徒テストデータ']
                
                # ヘッダー行のスタイル
                from openpyxl.styles import Font, PatternFill, Alignment
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # 列幅の自動調整
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            year_str = year if year and year != 'all' else 'all'
            period_str = period if period and period != 'all' else 'all'
            response['Content-Disposition'] = f'attachment; filename="student_comprehensive_data_{year_str}_{period_str}_{timestamp}.xlsx"'
            return response
        
        # GET リクエストの場合はエクスポート選択画面を表示
        # 年度と期間の選択肢を取得
        years = [{'value': 'all', 'label': '全年度'}]
        year_list = TestSchedule.objects.values_list('year', flat=True).distinct().order_by('-year')
        for year in year_list:
            years.append({'value': str(year), 'label': f'{year}年度'})
        
        periods = [
            {'value': 'all', 'label': '全期間'},
            {'value': 'spring', 'label': '春期'},
            {'value': 'summer', 'label': '夏期'}, 
            {'value': 'winter', 'label': '冬期'},
        ]
        
        context = {
            'title': '生徒データエクスポート（包括版）',
            'opts': Student._meta,
            'student_count': Student.objects.count(),
            'years': years,
            'periods': periods,
        }
        return render(request, 'admin/student_export.html', context)
    
    fieldsets = (
        ('基本情報', {
            'fields': ('student_id', 'name', 'grade', 'get_grade_display_readonly', 'classroom', 'is_active'),
            'description': '学年は数値で入力してください（小学1年生=1、小学6年生=6、中学1年生=7、中学3年生=9）'
        }),
        ('システム情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )


# 標準管理サイトにモデルを登録
admin.site.register(User, CustomUserAdmin)
admin.site.register(School, SchoolAdmin)
admin.site.register(Classroom, ClassroomAdmin)
admin.site.register(Student, StudentAdmin)

# テスト管理の追加（インポートは上部で済み）

class TestScheduleAdmin(admin.ModelAdmin):
    list_display = ('year', 'period', 'planned_date', 'actual_date', 'deadline_at', 'is_active')
    list_filter = ('year', 'period', 'is_active')
    search_fields = ('year', 'period')
    ordering = ('-year', 'period')

class QuestionInlineForGroup(admin.TabularInline):
    """大問内の小問用インライン"""
    model = Question
    extra = 2
    fields = ('question_number', 'content', 'max_score')
    ordering = ('question_number',)
    verbose_name = '小問'
    verbose_name_plural = '小問一覧'
    help_texts = {
        'question_number': '小問番号（1, 2, 3...）',
        'content': '問題文（簡潔に記入）',
        'max_score': '小問の配点'
    }

class QuestionGroupInline(admin.StackedInline):
    model = QuestionGroup
    extra = 3  # Show 3 empty forms by default
    min_num = 1  # Require at least 1 question group
    max_num = 10  # Allow up to 10 question groups
    fields = ('group_number', 'title', 'max_score')
    help_texts = {
        'group_number': '大問の番号を入力してください（例：1, 2, 3...）',
        'title': '大問のタイトルを入力してください（例：漢字の読み、文章読解、計算問題）',
        'max_score': 'この大問の満点を入力してください（例：20点、15点など）'
    }
    verbose_name = '大問'
    verbose_name_plural = '大問設定（必須）'
    can_delete = True
    show_change_link = False
    template = 'admin/edit_inline/stacked.html'
    
    def get_extra(self, request, obj=None, **kwargs):
        """If object exists and has question groups, show fewer extra forms"""
        if obj and obj.question_groups.exists():
            return 1
        return 3
    
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
        js = ('admin/js/test_definition.js',)
    
class TestDefinitionForm(forms.ModelForm):
    class Meta:
        model = TestDefinition
        fields = '__all__'
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # If grade_level is selected, filter subjects accordingly
        if 'grade_level' in self.data:
            try:
                grade_level = self.data['grade_level']
                self.fields['subject'].choices = TestDefinition.get_subjects_for_grade(grade_level)
            except (ValueError, TypeError):
                pass
        elif self.instance and self.instance.pk:
            # For editing existing objects
            self.fields['subject'].choices = TestDefinition.get_subjects_for_grade(self.instance.grade_level)

class TestDefinitionAdmin(admin.ModelAdmin):
    form = TestDefinitionForm
    list_display = ('get_test_display', 'grade_level', 'subject', 'max_score', 'get_question_groups_count', 'get_pdf_status', 'is_active', 'created_at')
    list_filter = ('grade_level', 'subject', 'is_active', 'schedule__year', 'schedule__period')
    search_fields = ('schedule__year',)
    inlines = [QuestionGroupInline]
    actions = ['create_test_template']
    
    def get_test_display(self, obj):
        return f"{obj.schedule.year}年度 {obj.schedule.get_period_display()}"
    get_test_display.short_description = '実施時期'
    get_test_display.admin_order_field = 'schedule__year'
    
    def get_question_groups_count(self, obj):
        return obj.question_groups.count()
    get_question_groups_count.short_description = '大問数'
    
    def get_pdf_status(self, obj):
        status = []
        if obj.question_pdf:
            status.append('問題PDF')
        if obj.answer_pdf:
            status.append('解答PDF')
        return ' / '.join(status) if status else '未登録'
    get_pdf_status.short_description = 'PDF状況'
    
    def create_test_template(self, request, queryset):
        """テストテンプレート作成"""
        self.message_user(request, "テストテンプレート機能は今後実装予定です。現在は手動で大問・小問を作成してください。")
    create_test_template.short_description = "テストテンプレートを作成"
    
    fieldsets = (
        ('テスト基本情報', {
            'fields': ('schedule', 'grade_level', 'subject', 'max_score'),
            'description': '実施するテストの基本情報を設定します。対象学年を選択すると、対応する科目が表示されます。満点は下記の大問の合計点と一致させてください。'
        }),
        ('PDFファイル', {
            'fields': ('question_pdf', 'answer_pdf'),
            'description': '問題用紙と解答用紙のPDFファイルをアップロードします。',
            'classes': ('collapse',)
        }),
        ('設定', {
            'fields': ('is_active',),
            'classes': ('collapse',)
        }),
    )
    
    def save_model(self, request, obj, form, change):
        """Save the test definition and update max_score if needed"""
        super().save_model(request, obj, form, change)
        # After saving, update max_score to match question groups total
        if obj.question_groups.exists():
            total_score = sum(group.max_score for group in obj.question_groups.all())
            if total_score != obj.max_score:
                obj.max_score = total_score
                obj.save(update_fields=['max_score'])
    
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('schedule')

class QuestionInline(admin.TabularInline):
    model = Question
    extra = 1
    fields = ('question_number', 'content', 'max_score')
    ordering = ('question_number',)

class AnswerKeyInline(admin.StackedInline):
    model = AnswerKey
    extra = 0
    fields = ('correct_answer', 'explanation')

class QuestionGroupAdmin(admin.ModelAdmin):
    list_display = ('get_test_info', 'group_number', 'title', 'max_score', 'get_question_count')
    list_filter = ('test__subject', 'test__schedule__year', 'test__schedule__period')
    search_fields = ('title', 'test__name')
    inlines = [QuestionInlineForGroup]
    ordering = ('test', 'group_number')
    
    fieldsets = (
        ('大問基本情報', {
            'fields': ('test', 'group_number', 'title', 'max_score'),
            'description': '大問の基本情報を設定します。保存後、下記で小問を追加できます。'
        }),
    )
    
    def get_test_info(self, obj):
        return f"{obj.test.schedule.year}年度 {obj.test.schedule.get_period_display()} {obj.test.get_subject_display()}"
    get_test_info.short_description = 'テスト情報'
    get_test_info.admin_order_field = 'test__schedule__year'
    
    def get_question_count(self, obj):
        return obj.questions.count()
    get_question_count.short_description = '小問数'
    
    fieldsets = (
        ('大問設定', {
            'fields': ('test', 'group_number', 'title', 'max_score'),
            'description': '大問の基本設定を行います。保存後、下記で小問を追加できます。'
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('test', 'test__schedule')

class QuestionAdmin(admin.ModelAdmin):
    list_display = ('get_question_info', 'question_number', 'content_preview', 'max_score', 'has_answer_key')
    list_filter = ('group__test__subject', 'group__test__schedule__year', 'group__test__schedule__period')
    search_fields = ('content', 'group__title', 'group__test__name')
    inlines = [AnswerKeyInline]
    ordering = ('group', 'question_number')
    
    def get_question_info(self, obj):
        return f"大問{obj.group.group_number} ({obj.group.title})"
    get_question_info.short_description = '所属大問'
    get_question_info.admin_order_field = 'group__group_number'
    
    def content_preview(self, obj):
        return obj.content[:50] + '...' if len(obj.content) > 50 else obj.content
    content_preview.short_description = '問題内容'
    
    def has_answer_key(self, obj):
        return hasattr(obj, 'answer_key')
    has_answer_key.boolean = True
    has_answer_key.short_description = '解答あり'
    
    fieldsets = (
        ('小問設定', {
            'fields': ('group', 'question_number', 'content', 'max_score'),
            'description': '小問の詳細を設定します。解答は下記で設定できます。'
        }),
    )
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('group', 'group__test')

class ScoreAdmin(admin.ModelAdmin):
    list_display = ('student', 'test', 'get_question_group_info', 'score', 'get_max_score', 'created_at')
    list_filter = ('test__subject', 'test__schedule__year', 'created_at')
    search_fields = ('student__name', 'student__student_id')
    actions = ['export_scores']
    
    def get_question_group_info(self, obj):
        return f"大問{obj.question_group.group_number} ({obj.question_group.title})"
    get_question_group_info.short_description = '大問'
    
    def get_max_score(self, obj):
        return f"{obj.score}/{obj.question_group.max_score}"
    get_max_score.short_description = '得点/満点'
    
    def export_scores(self, request, queryset):
        """得点データをエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        data = []
        for score in queryset:
            data.append({
                '塾ID': score.student.classroom.school.school_id,
                '塾名': score.student.classroom.school.name,
                '教室ID': score.student.classroom.classroom_id,
                '教室名': score.student.classroom.name,
                '生徒ID': score.student.student_id,
                '生徒名': score.student.name,
                'テスト': str(score.test),
                '科目': score.test.get_subject_display(),
                '大問番号': score.question_group.group_number,
                '大問名': score.question_group.title,
                '得点': score.score,
                '満点': score.question_group.max_score,
                '登録日': score.created_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='得点データ')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="scores_export_{timestamp}.xlsx"'
        return response
    
    export_scores.short_description = "得点データをエクスポート"
    
    def get_urls(self):
        """カスタムURLを追加"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('import-scores/', self.admin_site.admin_view(self.import_scores_view), name='scores_score_import'),
        ]
        return custom_urls + urls
    
    def import_scores_view(self, request):
        """得点インポートビュー（年度・時期・科目選択対応）"""
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from scores.utils import import_scores_from_excel, generate_score_template, get_available_tests
        from django.http import HttpResponse
        import tempfile
        import pandas as pd
        
        available_tests = get_available_tests()
        
        if request.method == 'POST':
            if 'excel_file' in request.FILES:
                file = request.FILES['excel_file']
                year = request.POST.get('year')
                period = request.POST.get('period')
                subject = request.POST.get('subject')
                
                if not year or not period or not subject:
                    messages.error(request, "年度、時期、学校種別をすべて選択してください。")
                    return redirect(request.get_full_path())
                
                # 一時ファイルに保存
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    for chunk in file.chunks():
                        tmp_file.write(chunk)
                    tmp_file_path = tmp_file.name
                
                try:
                    # インポート実行
                    result = import_scores_from_excel(tmp_file_path, int(year), period, subject)
                    
                    if result['success']:
                        success_msg = f"得点データのインポートが完了しました。\n"
                        success_msg += f"対象テスト: {result['test_info']}\n"
                        success_msg += f"新規作成: {result['created_scores']}件\n"
                        success_msg += f"更新: {result['updated_scores']}件"
                        messages.success(request, success_msg)
                        
                        if result['errors']:
                            error_msg = "以下のエラーがありました:\n" + "\n".join(result['errors'])
                            messages.warning(request, error_msg)
                    else:
                        messages.error(request, f"インポートに失敗しました: {result['error']}")
                        
                finally:
                    import os
                    os.unlink(tmp_file_path)
                
                return redirect('/admin/scores/score/')
            
            elif 'download_template' in request.POST:
                year = request.POST.get('year')
                period = request.POST.get('period')
                subject = request.POST.get('subject')
                
                if not year or not period or not subject:
                    messages.error(request, "年度、時期、学校種別をすべて選択してください。")
                    return redirect(request.get_full_path())
                
                try:
                    # テンプレート生成
                    df, structure = generate_score_template(int(year), period, subject)
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                        df.to_excel(tmp_file.name, index=False)
                        tmp_file_path = tmp_file.name
                    
                    try:
                        with open(tmp_file_path, 'rb') as f:
                            response = HttpResponse(
                                f.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
                            period_display = {'spring': '春季', 'summer': '夏季', 'winter': '冬季'}.get(period, period)
                            filename = f"score_template_{year}_{period_display}_{subject}.xlsx"
                            response['Content-Disposition'] = f'attachment; filename="{filename}"'
                            return response
                    finally:
                        import os
                        os.unlink(tmp_file_path)
                        
                except Exception as e:
                    messages.error(request, f"テンプレート生成に失敗しました: {str(e)}")
                    return redirect(request.get_full_path())
        
        context = {
            'title': '得点データ一括インポート',
            'opts': Score._meta,
            'available_tests': available_tests,
        }
        return render(request, 'admin/score_import.html', context)

class TestResultAdmin(admin.ModelAdmin):
    list_display = ('student', 'test', 'total_score', 'correct_rate', 'school_rank', 'national_rank')
    list_filter = ('test__subject', 'test__schedule__year', 'test__schedule__period', 'student__grade')
    search_fields = ('student__name', 'student__student_id')
    actions = ['export_test_results']
    readonly_fields = ('get_question_scores',)
    fields = ('student', 'test', 'total_score', 'correct_rate', 'school_rank', 'national_rank', 'comment', 'get_question_scores')
    
    def get_question_scores(self, obj):
        """大問ごとの得点を表示"""
        if not obj:
            return "データなし"
        
        from scores.models import Score
        scores = Score.objects.filter(student=obj.student, test=obj.test).order_by('question_group__group_number')
        
        if not scores.exists():
            return "大問ごとの得点データがありません"
        
        result = []
        for score in scores:
            if score.question_group:
                result.append(f"大問{score.question_group.group_number}: {score.score}点/{score.question_group.max_score}点")
            else:
                result.append(f"未分類: {score.score}点")
        
        return "\n".join(result)
    
    get_question_scores.short_description = "大問ごとの得点"
    
    def export_test_results(self, request, queryset):
        """テスト結果をエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        data = []
        for result in queryset.select_related('student', 'test', 'student__classroom__school'):
            data.append({
                '塾ID': result.student.classroom.school.school_id,
                '塾名': result.student.classroom.school.name,
                '教室ID': result.student.classroom.classroom_id,
                '教室名': result.student.classroom.name,
                '生徒ID': result.student.student_id,
                '生徒名': result.student.name,
                '学年': result.student.grade,
                'テスト': str(result.test),
                '年度': result.test.schedule.year,
                '時期': result.test.schedule.get_period_display(),
                '科目': result.test.get_subject_display(),
                '合計点': result.total_score,
                '正答率': f"{result.correct_rate:.1f}%",
                '塾内順位': f"{result.school_rank}/{result.school_total_students}",
                '全体順位': f"{result.national_rank}/{result.national_total_students}",
                'コメント': result.comment,
                '更新日': result.updated_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='テスト結果')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="test_results_{timestamp}.xlsx"'
        return response
    
    export_test_results.short_description = "テスト結果をエクスポート"
    
    def get_urls(self):
        """カスタムURLを追加"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path('summary/', self.admin_site.admin_view(self.test_result_summary_view), name='tests_testresult_summary'),
        ]
        return custom_urls + urls
    
    def test_result_summary_view(self, request):
        """テスト結果集計画面（年度・時期のみで集計、結果をDB保存）"""
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from scores.utils import get_available_tests, calculate_and_save_test_summary, get_test_summary, calculate_and_save_test_summary_by_school_type, get_test_summary_by_school_type
        
        available_tests = get_available_tests()
        
        # 年度・時期・科目の選択肢を生成
        years = sorted(list(set([test['year'] for test in available_tests])), reverse=True)
        # テストスケジュールモデルの期間を使用
        from tests.models import TestSchedule
        periods = TestSchedule.PERIODS
        # 学校種別の選択肢（テンプレートで直接定義するため不要）
        school_types = [
            ('elementary', '小学生'),
            ('middle', '中学生'),
        ]
        
        summary_data = None
        action = request.POST.get('action', '')
        
        # 選択された値を保持
        selected_year = ''
        selected_period = ''
        selected_school_type = ''
        
        if request.method == 'POST':
            year = request.POST.get('year')
            period = request.POST.get('period')
            school_type = request.POST.get('school_type')
            
            # 選択された値を保持（文字列として）
            selected_year = str(year) if year else ''
            selected_period = period or ''
            selected_school_type = school_type or ''
            
            if year and period and school_type:
                if action == 'calculate':
                    # 集計実行
                    try:
                        result = calculate_and_save_test_summary_by_school_type(int(year), period, school_type)
                        
                        if result['success']:
                            messages.success(request, 
                                f"集計が完了しました。対象: {result['total_students']}名、塾数: {result['schools_count']}")
                            # 集計後、結果を表示
                            summary_result = get_test_summary_by_school_type(int(year), period, school_type)
                            if summary_result['success']:
                                summary_data = summary_result
                        else:
                            messages.error(request, f"集計に失敗しました: {result['error']}")
                            
                    except Exception as e:
                        messages.error(request, f"集計エラー: {str(e)}")
                        
                elif action == 'view':
                    # 既存の集計結果を表示
                    try:
                        result = get_test_summary_by_school_type(int(year), period, school_type)
                        if result['success']:
                            summary_data = result
                        else:
                            messages.error(request, f"集計結果が見つかりません: {result['error']}")
                    except Exception as e:
                        messages.error(request, f"表示エラー: {str(e)}")
            else:
                messages.error(request, "年度、時期、学校種別をすべて選択してください。")
        
        context = {
            'title': 'テスト結果集計',
            'opts': TestResult._meta,
            'years': years,
            'periods': periods,
            'school_types': school_types,
            'summary_data': summary_data,
            'selected_year': selected_year,
            'selected_period': selected_period,
            'selected_school_type': selected_school_type,
        }
        return render(request, 'admin/test_result_summary.html', context)

class CommentTemplateAdmin(admin.ModelAdmin):
    list_display = ('school', 'subject', 'score_range_min', 'score_range_max', 'is_default', 'is_active')
    list_filter = ('subject', 'is_default', 'is_active', 'school')
    search_fields = ('template_text', 'school__name')


class TestSummaryAdmin(admin.ModelAdmin):
    list_display = ('__str__', 'total_students', 'average_score', 'average_correct_rate', 'calculated_at')
    list_filter = ('year', 'period', 'subject', 'calculated_at')
    search_fields = ('year', 'subject')
    readonly_fields = ('calculated_at', 'updated_at')
    
    fieldsets = (
        ('基本情報', {
            'fields': ('test', 'year', 'period', 'subject', 'max_score')
        }),
        ('統計情報', {
            'fields': ('total_students', 'average_score', 'average_correct_rate')
        }),
        ('詳細データ', {
            'fields': ('grade_statistics', 'school_statistics'),
            'classes': ('collapse',)
        }),
        ('日時情報', {
            'fields': ('calculated_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

class SchoolTestSummaryAdmin(admin.ModelAdmin):
    list_display = ('test_summary', 'school', 'student_count', 'average_score', 'rank_among_schools')
    list_filter = ('test_summary__year', 'test_summary__period', 'test_summary__subject', 'rank_among_schools')
    search_fields = ('school__name', 'school__school_id')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('基本情報', {
            'fields': ('test_summary', 'school', 'rank_among_schools')
        }),
        ('統計情報', {
            'fields': ('student_count', 'average_score', 'average_correct_rate')
        }),
        ('詳細データ', {
            'fields': ('grade_details',),
            'classes': ('collapse',)
        }),
        ('日時情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

# 解答管理
class AnswerKeyAdmin(admin.ModelAdmin):
    list_display = ('question', 'get_question_content', 'correct_answer_preview')
    list_filter = ('question__group__test__subject', 'question__group__test__schedule__year')
    search_fields = ('question__content', 'correct_answer')
    
    def get_question_content(self, obj):
        return obj.question.content[:50] + '...' if len(obj.question.content) > 50 else obj.question.content
    get_question_content.short_description = '問題内容'
    
    def correct_answer_preview(self, obj):
        return obj.correct_answer[:30] + '...' if len(obj.correct_answer) > 30 else obj.correct_answer
    correct_answer_preview.short_description = '正解'
    
    fieldsets = (
        ('解答情報', {
            'fields': ('question', 'correct_answer', 'explanation')
        }),
    )

# テスト関連の登録
admin.site.register(TestSchedule, TestScheduleAdmin)
# TestDefinitionは scores.admin で拡張版を登録
# admin.site.register(TestDefinition, TestDefinitionAdmin)
admin.site.register(QuestionGroup, QuestionGroupAdmin)
admin.site.register(Question, QuestionAdmin)
admin.site.register(AnswerKey, AnswerKeyAdmin)
admin.site.register(Score, ScoreAdmin)
admin.site.register(TestResult, TestResultAdmin)
admin.site.register(CommentTemplate, CommentTemplateAdmin)
admin.site.register(TestSummary, TestSummaryAdmin)
admin.site.register(SchoolTestSummary, SchoolTestSummaryAdmin)

# 会員種別・課金管理
class MembershipTypeAdmin(admin.ModelAdmin):
    list_display = ('name', 'type_code', 'price_per_student', 'is_active', 'created_at')
    list_filter = ('is_active', 'created_at')
    search_fields = ('name', 'type_code')
    readonly_fields = ('created_at', 'updated_at')
    
    fieldsets = (
        ('基本情報', {
            'fields': ('type_code', 'name', 'description', 'price_per_student', 'is_active')
        }),
        ('システム情報', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

# BillingReportAdminクラスは廃止（教室ベース課金レポートの代わりに塾ベースを使用）
# class BillingReportAdmin(admin.ModelAdmin):
#     list_display = ('classroom', 'year', 'period', 'billed_students', 'price_per_student', 'total_amount', 'generated_at')
#     list_filter = ('year', 'period', 'generated_at')
#     search_fields = ('classroom__name', 'classroom__school__name')
#     readonly_fields = ('generated_at', 'updated_at')
#     actions = ['export_billing_reports', 'export_all_billing_data']
    
    def export_billing_reports(self, request, queryset):
        """課金レポートをエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime
        
        data = []
        for report in queryset.select_related('classroom', 'classroom__school'):
            data.append({
                '塾ID': report.classroom.school.school_id,
                '塾名': report.classroom.school.name,
                '教室ID': report.classroom.classroom_id,
                '教室名': report.classroom.name,
                '会員種別': report.classroom.school.get_membership_type_display(),
                '年度': report.year,
                '期': report.get_period_display(),
                '総生徒数': report.total_students,
                '課金対象生徒数': report.billed_students,
                '単価（円）': report.price_per_student,
                '合計金額（円）': report.total_amount,
                '生成日': report.generated_at.strftime('%Y-%m-%d %H:%M'),
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='課金レポート')
        
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="billing_reports_{timestamp}.xlsx"'
        return response
    
    export_billing_reports.short_description = "課金レポートをエクスポート"
    
    def export_all_billing_data(self, request, queryset):
        """年度・期間を選択して請求情報を一括エクスポート"""
        from django.shortcuts import render, redirect
        from django.contrib import messages
        from django.http import HttpResponse
        from django.db.models import Q
        import io
        import pandas as pd
        from datetime import datetime
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # POSTリクエストの場合、エクスポート実行
        if request.method == 'POST':
            year = request.POST.get('year')
            period = request.POST.get('period')
            export_format = request.POST.get('export_format', 'excel')
            
            if not year or not period:
                messages.error(request, '年度と期間を選択してください。')
                return redirect(request.get_full_path())
            
            try:
                # 指定された年度・期間の得点データから出席済み生徒を取得
                from scores.models import Score, TestResult
                from tests.models import TestDefinition, TestSchedule
                from django.db.models import Q, Count, Distinct
                from collections import defaultdict
                
                # 指定年度・期間のテストを取得
                test_schedules = TestSchedule.objects.filter(year=year, period=period)
                test_definitions = TestDefinition.objects.filter(schedule__in=test_schedules)
                
                if not test_definitions.exists():
                    messages.error(request, f'指定された条件（{year}年度 {period}）のテストが見つかりません。')
                    return redirect(request.get_full_path())
                
                # 出席済み（attendance=True）の得点データを取得
                attended_scores = Score.objects.filter(
                    test__in=test_definitions,
                    attendance=True
                ).select_related(
                    'student',
                    'student__classroom',
                    'student__classroom__school',
                    'test',
                    'test__schedule'
                )  # SQLiteではdistinct(fields)をサポートしていないため、Pythonで重複排除
                
                if not attended_scores.exists():
                    messages.error(request, '指定された条件の出席済み生徒データが見つかりません。')
                    return redirect(request.get_full_path())
                
                # 塾ごとに出席生徒をグループ化（生徒・テスト組み合わせの重複排除）
                school_data = defaultdict(lambda: {
                    'school': None,
                    'students': set(),
                    'student_details': [],
                    'tests_taken': defaultdict(set)
                })
                
                # 生徒・テスト組み合わせの重複排除用
                processed_combinations = set()
                
                for score in attended_scores:
                    school = score.student.classroom.school
                    school_id = school.school_id
                    
                    # 生徒・テストの組み合わせをチェック（重複排除）
                    combination_key = (score.student.student_id, score.test.id)
                    if combination_key in processed_combinations:
                        continue
                    processed_combinations.add(combination_key)
                    
                    if school_data[school_id]['school'] is None:
                        school_data[school_id]['school'] = school
                    
                    # 生徒を追加（重複排除）
                    student_key = (score.student.student_id, score.student.name)
                    school_data[school_id]['students'].add(student_key)
                    
                    # テスト受験記録を追加
                    test_key = f"{score.test.get_subject_display()}"
                    school_data[school_id]['tests_taken'][student_key].add(test_key)
                
                # エクスポートデータを準備
                export_data = []
                total_amount_all = 0
                total_students_all = 0
                
                # 会員種別ごとの集計用
                membership_summary = defaultdict(lambda: {
                    'schools': set(),
                    'students': 0,
                    'amount': 0,
                    'price': 0
                })
                
                for school_id, data in school_data.items():
                    school = data['school']
                    student_count = len(data['students'])
                    
                    if student_count == 0:
                        continue
                    
                    membership_type = school.get_membership_type_display()
                    price_per_student = school.get_price_per_student()
                    total_amount = student_count * price_per_student
                    
                    # 会員種別ごとの集計
                    membership_summary[membership_type]['schools'].add(school_id)
                    membership_summary[membership_type]['students'] += student_count
                    membership_summary[membership_type]['amount'] += total_amount
                    membership_summary[membership_type]['price'] = price_per_student
                    
                    # 生徒詳細リストを作成
                    student_details_list = []
                    for student_id, student_name in data['students']:
                        tests_list = list(data['tests_taken'][(student_id, student_name)])
                        student_details_list.append(f"{student_name} (ID:{student_id}) - {', '.join(tests_list)}")
                    
                    export_data.append({
                        '塾ID': school.school_id,
                        '塾名': school.name,
                        '会員種別': membership_type,
                        '担当者': school.contact_person or '',
                        'メールアドレス': school.email or '',
                        '電話番号': school.phone or '',
                        '住所': school.address or '',
                        '年度': year,
                        '期': {'spring': '春期', 'summer': '夏期', 'winter': '冬期'}.get(period, period),
                        '出席生徒数': student_count,
                        '単価（円）': price_per_student,
                        '合計金額（円）': total_amount,
                        '受験生徒詳細': '\n'.join(student_details_list),
                    })
                    
                    total_amount_all += total_amount
                    total_students_all += student_count
                
                # 会員種別サマリーデータを作成
                summary_data = []
                for membership_type, data in membership_summary.items():
                    summary_data.append({
                        '会員種別': membership_type,
                        '塾数': len(data['schools']),
                        '出席生徒数': data['students'],
                        '単価（円）': data['price'],
                        '合計金額（円）': data['amount'],
                    })
                
                # 合計行を追加
                summary_data.append({
                    '会員種別': '【合計】',
                    '塾数': len(set([d['塾ID'] for d in export_data])),
                    '出席生徒数': total_students_all,
                    '単価（円）': '-',
                    '合計金額（円）': total_amount_all,
                })
                
                # データフレーム作成
                df_details = pd.DataFrame(export_data)
                df_summary = pd.DataFrame(summary_data)
                
                if export_format == 'csv':
                    # CSV形式でエクスポート
                    output = io.StringIO()
                    
                    # サマリー情報をCSVに含める
                    output.write(f"# 請求情報エクスポート - {year}年度 {period}\n")
                    output.write(f"# 生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    output.write("# \n")
                    output.write("# ========== サマリー情報 ==========\n")
                    df_summary.to_csv(output, index=False, encoding='utf-8-sig')
                    output.write("# \n")
                    output.write("# ========== 詳細情報 ==========\n")
                    df_details.to_csv(output, index=False, encoding='utf-8-sig')
                    
                    response = HttpResponse(
                        output.getvalue(),
                        content_type='text/csv; charset=utf-8-sig'
                    )
                    period_display = {'spring': '春期', 'summer': '夏期', 'winter': '冬期'}.get(period, period)
                    filename = f'billing_report_{year}_{period_display}.csv'
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    return response
                
                else:
                    # Excel形式でエクスポート
                    output = io.BytesIO()
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # サマリーシート
                        df_summary.to_excel(writer, sheet_name='サマリー', index=False)
                        
                        # 詳細データシート
                        df_details.to_excel(writer, sheet_name='詳細データ', index=False)
                        
                        # サマリーシートのスタイル
                        summary_ws = writer.sheets['サマリー']
                        
                        # ヘッダー行のスタイル
                        header_font = Font(bold=True, color='FFFFFF')
                        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                        
                        for cell in summary_ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')
                        
                        # 合計行のスタイル（最終行）
                        total_row = summary_ws.max_row
                        total_font = Font(bold=True, color='FFFFFF')
                        total_fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
                        
                        for cell in summary_ws[total_row]:
                            cell.font = total_font
                            cell.fill = total_fill
                            cell.alignment = Alignment(horizontal='center')
                        
                        # 列幅の自動調整
                        for column in summary_ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 30)
                            summary_ws.column_dimensions[column_letter].width = adjusted_width
                        
                        # 詳細データシートのスタイル
                        detail_ws = writer.sheets['詳細データ']
                        
                        # ヘッダー行のスタイル
                        for cell in detail_ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal='center')
                        
                        # 列幅の自動調整
                        for column in detail_ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 40)
                            detail_ws.column_dimensions[column_letter].width = adjusted_width
                    
                    output.seek(0)
                    response = HttpResponse(
                        output.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                    
                    period_display = {'spring': '春期', 'summer': '夏期', 'winter': '冬期'}.get(period, period)
                    filename = f'billing_report_{year}_{period_display}.xlsx'
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    
                    # 成功メッセージ
                    self.message_user(
                        request, 
                        f'請求情報をエクスポートしました: {year}年度 {period_display} - {len(export_data)}塾、{total_students_all}名、{total_amount_all:,}円'
                    )
                    return response
                    
            except Exception as e:
                messages.error(request, f'エクスポートエラー: {str(e)}')
                return redirect(request.get_full_path())
        
        # GETリクエストの場合、年度・期間選択画面を表示
        from tests.models import TestSchedule
        years = TestSchedule.objects.values_list('year', flat=True).distinct().order_by('-year')
        
        periods = [
            {'value': 'spring', 'label': '春期'},
            {'value': 'summer', 'label': '夏期'},
            {'value': 'winter', 'label': '冬期'},
        ]
        
        context = {
            'title': '請求情報一括エクスポート',
            'opts': self.model._meta,
            'years': years,
            'periods': periods,
            'action_name': 'export_all_billing_data',
        }
        
        return render(request, 'admin/export_billing_data.html', context)
    
#     export_all_billing_data.short_description = "📊 請求情報を一括エクスポート（年度・期間選択）"
    
    fieldsets = (
        ('基本情報', {
            'fields': ('classroom', 'year', 'period')
        }),
        ('集計情報', {
            'fields': ('total_students', 'billed_students', 'price_per_student', 'total_amount')
        }),
        ('詳細データ', {
            'fields': ('student_details',),
            'classes': ('collapse',)
        }),
        ('システム情報', {
            'fields': ('generated_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

admin.site.register(MembershipType, MembershipTypeAdmin)
# 教室ベースの課金レポートは非表示（塾ベースに統一）
# admin.site.register(BillingReport, BillingReportAdmin)

# BillingReportモデルは完全に廃止（コメントアウト）

# 課金レポート管理（塾ベース）
class SchoolBillingReportAdmin(admin.ModelAdmin):
    change_list_template = 'admin/classrooms/schoolbillingreport/change_list.html'
    list_display = ('school', 'year', 'period', 'total_classrooms', 'billed_students', 'price_per_student', 'total_amount', 'average_per_classroom', 'generated_at')
    list_filter = ('year', 'period', 'generated_at', 'school__membership_type')
    search_fields = ('school__name', 'school__school_id')
    readonly_fields = ('generated_at', 'updated_at')
    actions = ['regenerate_school_billing_reports', 'export_school_billing_reports', 'export_school_billing_data']

    # 手動での課金レポート作成を無効化（自動生成のみ）
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False  # 読み取り専用

    def changelist_view(self, request, extra_context=None):
        """一括追加機能付きの課金レポートリスト画面"""
        from django.urls import reverse

        extra_context = extra_context or {}
        extra_context['custom_add_url'] = reverse('bulk_add_school_billing')
        extra_context['show_bulk_add'] = True

        response = super().changelist_view(request, extra_context)

        if hasattr(response, 'context_data') and response.context_data:
            cl = response.context_data.get('cl')
            if cl is not None:
                queryset = cl.queryset
                from django.db.models import Sum, Count

                summary = queryset.aggregate(
                    total_schools=Count('pk', distinct=True),
                    total_classrooms=Sum('total_classrooms'),
                    total_students=Sum('billed_students'),
                    total_amount=Sum('total_amount'),
                )

                response.context_data['summary'] = {
                    'total_schools': summary.get('total_schools') or 0,
                    'total_classrooms': summary.get('total_classrooms') or 0,
                    'total_students': summary.get('total_students') or 0,
                    'total_amount': summary.get('total_amount') or 0,
                }

        return response


    def has_delete_permission(self, request, obj=None):
        return True  # 削除は可能

    def average_per_classroom(self, obj):
        avg = obj.get_average_per_classroom()
        return f"{avg:.0f}円/教室" if avg > 0 else "0円/教室"
    average_per_classroom.short_description = '教室あたり平均'

    def regenerate_school_billing_reports(self, request, queryset):
        """選択した塾の課金レポートを再生成"""
        from django.contrib import messages
        from classrooms.utils import generate_school_billing_report

        created = updated = skipped = errors = 0

        for report in queryset:
            try:
                result = generate_school_billing_report(
                    school=report.school,
                    year=report.year,
                    period=report.period,
                    force=True,
                )

                if result.get('created'):
                    created += 1
                elif result.get('updated'):
                    updated += 1
                else:
                    skipped += 1
            except Exception as exc:
                errors += 1
                messages.error(request, f"{report.school.name}（{report.year}年度{report.get_period_display()}）: {exc}")

        if created:
            messages.success(request, f"{created}件の塾別課金レポートを新規作成しました。")
        if updated:
            messages.success(request, f"{updated}件の塾別課金レポートを再計算しました。")
        if skipped and not (created or updated):
            messages.info(request, f"{skipped}件のレポートを処理しました（変更なし）。")
        if errors:
            messages.warning(request, f"{errors}件のレポートでエラーが発生しました。詳細はメッセージをご確認ください。")

    regenerate_school_billing_reports.short_description = "♻️ 選択した塾の課金レポートを再生成"

    def export_school_billing_reports(self, request, queryset):
        """塾別課金レポートをExcelエクスポート"""
        from django.http import HttpResponse
        import io
        import pandas as pd
        from datetime import datetime

        data = []
        for report in queryset:
            data.append({
                '塾ID': report.school.school_id,
                '塾名': report.school.name,
                '年度': report.year,
                '期': report.get_period_display(),
                '会員種別': report.school.get_membership_type_display(),
                '教室数': report.total_classrooms,
                '総生徒数': report.total_students,
                '課金対象生徒数': report.billed_students,
                '単価（円）': report.price_per_student,
                '合計金額（円）': report.total_amount,
                '教室あたり平均（円）': round(report.get_average_per_classroom()),
                '生成日': report.generated_at.strftime('%Y-%m-%d %H:%M'),
            })

        df = pd.DataFrame(data)
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='塾別課金レポート')

        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="school_billing_reports_{timestamp}.xlsx"'
        return response

    export_school_billing_reports.short_description = "💼 塾別課金レポートをExcelエクスポート"

    def export_school_billing_data(self, request, queryset):
        """年度・期間を選択して塾別請求情報を一括エクスポート"""
        from django.shortcuts import render
        from django.contrib import messages
        from classrooms.models import SchoolBillingReport
        import pandas as pd
        from django.http import HttpResponse
        import io
        from datetime import datetime

        # POSTリクエストの場合、エクスポート実行
        if request.method == 'POST':
            year = request.POST.get('year')
            period = request.POST.get('period')

            if not year or not period:
                messages.error(request, '年度と期間を選択してください')
                return

            try:
                year = int(year)
                reports = SchoolBillingReport.objects.filter(year=year, period=period)

                if not reports.exists():
                    messages.warning(request, f'{year}年度{period}期の課金レポートが見つかりません')
                    return

                # データを準備
                export_data = []
                total_amount_all = 0
                total_students_all = 0

                for report in reports:
                    export_data.append({
                        '塾ID': report.school.school_id,
                        '塾名': report.school.name,
                        '会員種別': report.school.get_membership_type_display(),
                        '教室数': report.total_classrooms,
                        '課金対象生徒数': report.billed_students,
                        '単価': report.price_per_student,
                        '合計金額': report.total_amount,
                        '教室あたり平均': round(report.get_average_per_classroom()),
                    })

                    total_amount_all += report.total_amount
                    total_students_all += report.billed_students

                period_display = {'spring': '春期', 'summer': '夏期', 'winter': '冬期'}.get(period, period)

                # Excelエクスポート
                df = pd.DataFrame(export_data)
                output = io.BytesIO()

                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=f'{year}年度{period_display}')

                output.seek(0)
                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

                filename = f'school_billing_report_{year}_{period_display}.xlsx'
                response['Content-Disposition'] = f'attachment; filename="{filename}"'

                messages.success(
                    request,
                    f'塾別請求情報をエクスポートしました: {year}年度 {period_display} - {len(export_data)}塾、{total_students_all}名、{total_amount_all:,}円'
                )

                return response

            except Exception as e:
                messages.error(request, f'エクスポート中にエラーが発生しました: {str(e)}')

        # GETリクエストの場合、フォーム表示
        return render(request, 'admin/export_billing_data.html', {
            'title': '塾別請求情報一括エクスポート',
            'opts': self.model._meta,
            'years': SchoolBillingReport.objects.values_list('year', flat=True).distinct().order_by('-year'),
            'periods': [('spring', '春期'), ('summer', '夏期'), ('winter', '冬期')],
            'action_name': 'export_school_billing_data',
        })

    export_school_billing_data.short_description = "📊 塾別請求情報を一括エクスポート（年度・期間選択）"

    fieldsets = (
        ('基本情報', {
            'fields': ('school', 'year', 'period')
        }),
        ('集計情報', {
            'fields': ('total_classrooms', 'total_students', 'billed_students', 'price_per_student', 'total_amount')
        }),
        ('詳細情報', {
            'fields': ('classroom_details', 'student_details'),
            'classes': ('collapse',)
        }),
        ('システム情報', {
            'fields': ('generated_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )

from classrooms.models import SchoolBillingReport
admin.site.register(SchoolBillingReport, SchoolBillingReportAdmin)

# Django管理画面のカスタマイズ
admin.site.site_header = '全国学力向上テスト 管理画面'
admin.site.site_title = '全国学力向上テスト'
admin.site.index_title = 'システム管理'
